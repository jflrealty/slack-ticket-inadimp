from database import SessionLocal
from models import OrdemServicoServicos
from datetime import datetime
import os
import io
import csv
import urllib.request
import unicodedata
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from slack_sdk import WebClient
from fpdf import FPDF

client_slack = WebClient(token=os.getenv("SLACK_BOT_TOKEN"))

# 🔧 Utilitário para nomes do Slack
def get_nome_slack(user_id):
    if not user_id or not user_id.startswith("U"):
        return user_id
    try:
        info = client_slack.users_info(user=user_id)
        return info["user"]["real_name"]
    except Exception as e:
        print(f"Erro ao buscar nome real de {user_id}: {e}")
        return user_id

# 🔧 Limpa texto para evitar erros de encoding (PDF e outros)
def limpar_texto_pdf(texto):
    if not texto:
        return ""
    texto = texto.replace("–", "-").replace("—", "-")
    texto = texto.replace("“", '"').replace("”", '"')
    texto = texto.replace("‘", "'").replace("’", "'")
    texto = unicodedata.normalize("NFKD", texto).encode("latin-1", "ignore").decode("latin-1")
    return texto

# 📋 Modal Serviços
def montar_blocos_modal_servicos():
    return [
        {
            "type": "input",
            "block_id": "tipo_ticket",
            "element": {
                "type": "static_select",
                "action_id": "value",
                "placeholder": {"type": "plain_text", "text": "Escolha"},
                "options": [
                    {"text": {"type": "plain_text", "text": "Suspensão de Serviços"}, "value": "Suspensão de Serviços"},
                    {"text": {"type": "plain_text", "text": "Restabelecimento de Serviços"}, "value": "Restabelecimento de Serviços"}
                ]
            },
            "label": {"type": "plain_text", "text": "Tipo de Ticket"}
        },
        {
            "type": "input",
            "block_id": "locatario",
            "element": {"type": "plain_text_input", "action_id": "value"},
            "label": {"type": "plain_text", "text": "Locatário"}
        },
        {
            "type": "input",
            "block_id": "empreendimento_unidade",
            "element": {"type": "plain_text_input", "action_id": "value"},
            "label": {"type": "plain_text", "text": "Empreendimento e Unidade"}
        },
        {
            "type": "input",
            "block_id": "responsavel",
            "element": {
                "type": "static_select",
                "action_id": "value",
                "placeholder": {"type": "plain_text", "text": "Escolha o responsável"},
                "options": [
                    {"text": {"type": "plain_text", "text": "Alef Nunes"}, "value": "U07DN49NT6V"},
                    {"text": {"type": "plain_text", "text": "Braullio Reis"}, "value": "U082F92585P"},
                    {"text": {"type": "plain_text", "text": "Gabriela Casão"}, "value": "U06U8AG164R"},
                    {"text": {"type": "plain_text", "text": "Jelifer Neves"}, "value": "U085ME3BYFP"},
                    {"text": {"type": "plain_text", "text": "Juliana"}, "value": "U08SMM12TM0"},
                    {"text": {"type": "plain_text", "text": "Marina Macena"}, "value": "U06UKKKNJTG"},
                    {"text": {"type": "plain_text", "text": "Marta Cabral"}, "value": "U07S6RFBGE6"},
                    {"text": {"type": "plain_text", "text": "Rafaela Assis"}, "value": "U092WBQKE11"},
                    {"text": {"type": "plain_text", "text": "Rafaela Oh"}, "value": "U08T7A8RWH4"},
                    {"text": {"type": "plain_text", "text": "Recepção AVNU"}, "value": "U083KH7R0AY"},
                    {"text": {"type": "plain_text", "text": "Recepcao Jml747"}, "value": "U06TVGUV119"},
                    {"text": {"type": "plain_text", "text": "Recepção Residencial - JFL125"}, "value": "U06TVLDAG9Y"},
                    {"text": {"type": "plain_text", "text": "Recepcao Vhouse"}, "value": "U06U8AF70M7"},
                    {"text": {"type": "plain_text", "text": "Recepcao Vo699"}, "value": "U06TY4N7JKE"},
                    {"text": {"type": "plain_text", "text": "Thaiane Leoni"}, "value": "U06TVLDE7AN"}
                ]
            },
            "label": {"type": "plain_text", "text": "Responsável"}
        }
    ]
# 💾 Criar OS
def criar_ordem_servico_servicos(data, thread_ts=None, canal_id=None):
    db = SessionLocal()
    try:
        chamado = OrdemServicoServicos(
            tipo_ticket=data["tipo_ticket"],
            locatario=data["locatario"],
            empreendimento_unidade=data["empreendimento_unidade"],
            responsavel=data["responsavel"],
            status="aberto",
            criado_em=datetime.utcnow(),
            thread_ts=thread_ts,
            canal_id=canal_id
        )
        db.add(chamado)
        db.commit()
    except Exception as e:
        print("❌ Erro ao salvar chamado:", e)
        db.rollback()
    finally:
        db.close()

# 🧾 Mensagem formatada
def formatar_mensagem_chamado_servicos(data, user_id):
    def limpar(valor):
        return valor if valor else "–"

    def formatar_mencao(slack_id):
        if not slack_id:
            return "–"
        return f"<@{slack_id}>"
    return (
        f"*Tipo:* {limpar(data.get('tipo_ticket'))}\n"
        f"*Locatário:* {limpar(data.get('locatario'))}\n"
        f"*Empreendimento e Unidade:* {limpar(data.get('empreendimento_unidade'))}\n"
        f"*Responsável:* {formatar_mencao(data.get('responsavel'))}\n"
        f"*Solicitante:* {formatar_mencao(user_id)}"
    )

# 🔄 Capturar chamado
def capturar_chamado(client, body):
    db = SessionLocal()
    try:
        user_id = body["user"]["id"]
        canal_id = body["channel"]["id"]
        ts = body["message"]["ts"]

        chamado = db.query(OrdemServicoServicos).filter_by(thread_ts=ts).first()
        if chamado:
            chamado.status = "em atendimento"
            chamado.responsavel = user_id
            chamado.atualizado_em = datetime.utcnow()
            db.commit()

            # Atualiza mensagem no Slack
            client.chat_update(
                channel=canal_id,
                ts=ts,
                text=f"🔄 Chamado capturado por <@{user_id}>",
                blocks=[
                    {
                        "type": "section",
                        "text": {"type": "mrkdwn", "text": f"🔄 Chamado capturado por <@{user_id}>"}
                    },
                    {
                        "type": "actions",
                        "elements": [
                            {"type": "button", "text": {"type": "plain_text", "text": "✅ Encerrar"}, "action_id": "finalizar_chamado"}
                        ]
                    }
                ]
            )
    except Exception as e:
        print("❌ Erro ao capturar chamado:", e)
        db.rollback()
    finally:
        db.close()

# ✅ Finalizar chamado
def finalizar_chamado(client, body):
    db = SessionLocal()
    try:
        user_id = body["user"]["id"]
        canal_id = body["channel"]["id"]
        ts = body["message"]["ts"]

        chamado = db.query(OrdemServicoServicos).filter_by(thread_ts=ts).first()
        if chamado:
            chamado.status = "encerrado"
            chamado.atualizado_em = datetime.utcnow()
            db.commit()

            client.chat_update(
                channel=canal_id,
                ts=ts,
                text=f"✅ Chamado finalizado por <@{user_id}>"
            )
    except Exception as e:
        print("❌ Erro ao finalizar chamado:", e)
        db.rollback()
    finally:
        db.close()

# 📋 Listar chamados por usuário
def listar_chamados_por_usuario_servicos(user_id):
    db = SessionLocal()
    chamados = db.query(OrdemServicoServicos)\
                 .filter_by(responsavel=user_id)\
                 .order_by(OrdemServicoServicos.criado_em.desc())\
                 .all()
    db.close()
    return chamados

# 📁 Buscar todos os chamados
def buscar_todos_chamados_servicos():
    db = SessionLocal()
    chamados = db.query(OrdemServicoServicos).order_by(OrdemServicoServicos.criado_em.desc()).all()
    db.close()
    return chamados

# 🗂️ Exportação (PDF, CSV, XLSX)
def gerar_pdf_chamados_servicos(chamados):
    pdf = FPDF()
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.add_page()

    LOGO_URL = "https://raw.githubusercontent.com/jflrealty/images/main/JFL_logotipo_completo.jpg"
    logo_path = "/tmp/logo_jfl.jpg"
    try:
        urllib.request.urlretrieve(LOGO_URL, logo_path)
        pdf.image(logo_path, x=10, y=8, w=60)
    except Exception as e:
        print(f"❌ Erro ao carregar logo: {e}")

    pdf.set_font("Arial", "B", 10)
    headers = ["ID", "Tipo", "Locatário", "Empreendimento", "Responsável", "Status", "Criado em"]
    col_widths = [10, 40, 40, 40, 40, 25, 30]

    for i, h in enumerate(headers):
        pdf.cell(col_widths[i], 8, limpar_texto_pdf(h), border=1, fill=True)
    pdf.ln()

    pdf.set_font("Arial", "", 9)
    for ch in chamados:
        row = [
            ch.id,
            ch.tipo_ticket,
            ch.locatario,
            ch.empreendimento_unidade,
            get_nome_slack(ch.responsavel),
            ch.status,
            ch.criado_em.strftime('%d/%m/%Y %H:%M')
        ]
        for i, c in enumerate(row):
            pdf.cell(col_widths[i], 8, limpar_texto_pdf(str(c)), border=1)
        pdf.ln()

    pdf_bytes = pdf.output(dest="S").encode("latin-1")
    return io.BytesIO(pdf_bytes)

def gerar_csv_chamados_servicos(chamados):
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["ID", "Tipo", "Locatário", "Empreendimento", "Responsável", "Status", "Criado em"])
    for ch in chamados:
        writer.writerow([
            ch.id,
            limpar_texto_pdf(ch.tipo_ticket),
            limpar_texto_pdf(ch.locatario),
            limpar_texto_pdf(ch.empreendimento_unidade),
            limpar_texto_pdf(get_nome_slack(ch.responsavel)),
            limpar_texto_pdf(ch.status),
            ch.criado_em.strftime('%d/%m/%Y %H:%M')
        ])
    output.seek(0)
    return io.BytesIO(output.read().encode("utf-8"))

def gerar_xlsx_chamados_servicos(chamados):
    wb = Workbook()
    ws = wb.active
    ws.title = "Chamados Serviços"

    colunas = ["ID", "Tipo", "Locatário", "Empreendimento", "Responsável", "Status", "Criado em"]
    ws.append(colunas)

    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    for c in chamados:
        ws.append([
            c.id,
            limpar_texto_pdf(c.tipo_ticket),
            limpar_texto_pdf(c.locatario),
            limpar_texto_pdf(c.empreendimento_unidade),
            limpar_texto_pdf(get_nome_slack(c.responsavel)),
            limpar_texto_pdf(c.status),
            c.criado_em.strftime('%d/%m/%Y %H:%M')
        ])

    for col in ws.columns:
        max_length = max(len(str(cell.value)) for cell in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = max(max_length + 2, 12)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output

# ✅ Exportar chamados
def exportar_chamados_servicos(data_inicio=None, data_fim=None, tipo="pdf"):
    chamados = buscar_todos_chamados_servicos()

    if data_inicio:
        chamados = [c for c in chamados if c.criado_em >= data_inicio]
    if data_fim:
        chamados = [c for c in chamados if c.criado_em <= data_fim]

    agora = datetime.now().strftime("%Y%m%d%H%M%S")

    return {
        "pdf": (f"chamados-servicos-{agora}.pdf", gerar_pdf_chamados_servicos(chamados)),
        "csv": (f"chamados-servicos-{agora}.csv", gerar_csv_chamados_servicos(chamados)),
        "xlsx": (f"chamados-servicos-{agora}.xlsx", gerar_xlsx_chamados_servicos(chamados))
    }

# 📦 Modal de exportação
def montar_blocos_exportacao_servicos():
    return [
        {
            "type": "input",
            "block_id": "tipo_arquivo",
            "label": {"type": "plain_text", "text": "Formato do Arquivo"},
            "element": {
                "type": "static_select",
                "action_id": "value",
                "placeholder": {"type": "plain_text", "text": "Escolha o formato"},
                "options": [
                    {"text": {"type": "plain_text", "text": "PDF"}, "value": "pdf"},
                    {"text": {"type": "plain_text", "text": "CSV"}, "value": "csv"},
                    {"text": {"type": "plain_text", "text": "Excel"}, "value": "xlsx"}
                ]
            }
        },
        {
            "type": "input",
            "block_id": "data_inicio",
            "label": {"type": "plain_text", "text": "Data Inicial"},
            "element": {"type": "datepicker", "action_id": "value"}
        },
        {
            "type": "input",
            "block_id": "data_fim",
            "label": {"type": "plain_text", "text": "Data Final"},
            "element": {"type": "datepicker", "action_id": "value"}
        }
    ]
